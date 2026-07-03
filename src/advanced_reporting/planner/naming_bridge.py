"""Bridge a ``CampaignPlan`` into the naming generator's input workbook (zero glue).

``write_plan_xlsx`` writes an xlsx with the ``Plan`` / ``Scheme`` / ``Settings`` sheets that
``naming/naming_generator.py:generate`` reads — the ``Plan`` headers are exactly ``PLAN_COLS``
and each row is one ``to_plan_rows()`` dict. ``openpyxl`` is imported lazily so it stays an
optional dependency (only needed when you actually hand a plan to the generator).
"""
from __future__ import annotations

from .schema import PLAN_COLS

# Mirror the generator's defaults (every Scheme field is a Plan column).
_DEFAULT_SCHEME = [
    ("Campaign", "market, channel, objective, audience_type", "_"),
    ("Ad Set", "audience_type, audience_detail, placement", "_"),
    ("Ad", "creative, format, size, version", "_"),
]
_DEFAULT_SETTINGS = [
    ("case", "asis"), ("max_name_length", "100"), ("max_combinations", "20000"),
    ("landing_url", "https://www.example.com"), ("utm_medium", "paid"),
]


def write_plan_xlsx(plan, path, *, scheme=None, settings=None):
    """Write ``plan`` as a generator-ready workbook at ``path``; return ``path``."""
    from openpyxl import Workbook  # lazy: optional dependency

    scheme = scheme or _DEFAULT_SCHEME
    settings = settings or _DEFAULT_SETTINGS
    wb = Workbook()
    wb.remove(wb.active)

    ws = wb.create_sheet("Plan")
    ws.append(PLAN_COLS)
    for row in plan.to_plan_rows():
        ws.append([row[c] for c in PLAN_COLS])

    sch = wb.create_sheet("Scheme")
    sch.append(["Level", "Fields (ordered, comma-separated)", "Delimiter"])
    for level, fields, delim in scheme:
        sch.append([level, fields, delim])

    sett = wb.create_sheet("Settings")
    sett.append(["Setting", "Value", "Notes"])
    for key, value in settings:
        sett.append([key, value, ""])

    wb.save(path)
    return path
