"""Naming-decode tests: generator names must round-trip back into their plan fields;
non-conforming names must land in the explicit "(unparsed)" bucket, never be guessed."""
from __future__ import annotations

import pandas as pd
import pytest

from advanced_reporting.ingestion import naming_decode as nd
from naming import naming_generator as gen


def test_round_trip_through_the_real_generator(tmp_path):
    """Names produced by the actual generator (its example plan) decode to the same fields."""
    template = tmp_path / "plan.xlsx"
    output = tmp_path / "out.xlsx"
    gen.build_template(template)
    _cols, records, warnings = gen.generate(template, output)
    assert records and not warnings
    for rec in records:
        ad_set = nd.decode_name(rec["Ad Set Name"])
        assert ad_set.kind == "ad_set"
        assert ad_set.audience_type == rec["audience_type"]
        assert ad_set.audience_detail == rec["audience_detail"]
        assert ad_set.placement == rec["placement"]
        ad = nd.decode_name(rec["Ad Name"])
        assert ad.kind == "ad"
        assert ad.creative == rec["creative"]
        assert ad.creative_format == rec["format"]


@pytest.mark.parametrize("name,kind,fields", [
    ("PROSPECT_LAL-1PCT_FEED", "ad_set",
     {"audience_type": "PROSPECT", "audience_detail": "LAL-1PCT"}),
    ("RETARGET_SITE-90D_REELS", "ad_set",
     {"audience_type": "RETARGET", "audience_detail": "SITE-90D"}),
    ("PROSPECT_KW-BRAND", "ad_set",       # placement omitted (generator skips empties)
     {"audience_type": "PROSPECT", "audience_detail": "KW-BRAND"}),
    ("BRANDHERO_VID_9x16_V1", "ad", {"creative": "BRANDHERO", "creative_format": "VID"}),
    ("BRANDHERO_VID_9x16", "ad",          # version omitted -> 3 tokens, still an Ad
     {"creative": "BRANDHERO", "creative_format": "VID"}),
    ("DEADLINE_STATIC_1x1_V2", "ad",
     {"creative": "DEADLINE", "creative_format": "STATIC"}),
])
def test_grammar_decodes(name, kind, fields):
    d = nd.decode_name(name)
    assert d.kind == kind
    for k, v in fields.items():
        assert getattr(d, k) == v


@pytest.mark.parametrize("name", [
    "Advantage+ broad (test)",        # spaces + illegal chars
    "MBA Degree - Broad (old)",
    "US_META_CONVERT_PROSPECT",       # campaign grammar: 4 tokens, no ad signals
    "one",                            # single token
    "a_b_c_d_e",                      # too many tokens
    "PROSPECT__FEED",                 # empty token
])
def test_non_conforming_names_land_in_unparsed(name):
    d = nd.decode_name(name)
    assert d.kind == "unparsed"
    assert d.audience_type == nd.UNPARSED
    # never guessed: every other field stays empty
    assert d.audience_detail == d.creative == d.creative_format == ""


# -- campaign grammar: the optional trailing `initiative` segment (backward compatible) --

@pytest.mark.parametrize("name,fields", [
    # every pre-initiative 4-segment campaign name must still decode unchanged
    ("US_META_CONVERT_PROSPECT",
     {"market": "US", "channel": "META", "objective": "CONVERT",
      "audience_type": "PROSPECT", "initiative": ""}),
    ("US_GOOGLE_AWARENESS_PROSPECT",
     {"market": "US", "channel": "GOOGLE", "objective": "AWARENESS",
      "audience_type": "PROSPECT", "initiative": ""}),
])
def test_campaign_decode_backward_compat(name, fields):
    d = nd.decode_campaign_name(name)
    assert d.kind == "campaign"
    for k, v in fields.items():
        assert getattr(d, k) == v


@pytest.mark.parametrize("name,initiative", [
    ("US_META_CONVERT_PROSPECT_SA", "SA"),
    ("US_LINKEDIN_CONSIDER_RETARGET_INTEL", "INTEL"),
    ("US_YOUTUBE_AWARENESS_PROSPECT_CYBER", "CYBER"),
])
def test_campaign_decode_with_initiative(name, initiative):
    d = nd.decode_campaign_name(name)
    assert d.kind == "campaign"
    assert d.initiative == initiative
    assert nd.decode_initiative(name) == initiative
    # segments 0-3 are read exactly as the 4-segment form
    assert (d.market, d.channel, d.objective, d.audience_type) == tuple(name.split("_")[:4])


@pytest.mark.parametrize("name", [
    "US_META_CONVERT",                 # too few segments
    "a_b_c_d_e_f",                     # too many segments
    "Advantage+ broad (test)",         # illegal characters
    "US_META_CONVERT_PROSPECT__SA",    # empty token
])
def test_campaign_decode_unparsed(name):
    d = nd.decode_campaign_name(name)
    assert d.kind == "unparsed"
    assert d.initiative == "" and d.market == ""


def test_campaign_decode_blank_and_helper():
    for blank in ("", "  ", None, float("nan")):
        d = nd.decode_campaign_name(blank)
        assert d.kind == "blank" and d.initiative == ""
    assert nd.decode_initiative("US_META_CONVERT_PROSPECT") == ""


def test_campaign_round_trip_through_generator(tmp_path):
    """The real generator composes the trailing initiative; decode recovers it, and blank
    initiatives still produce classic 4-segment names."""
    from openpyxl import load_workbook

    template = tmp_path / "plan.xlsx"
    output = tmp_path / "out.xlsx"
    gen.build_template(template)

    # Inject an initiative on the first two example rows; leave the rest blank.
    wb = load_workbook(template)
    ws = wb["Plan"]
    init_col = gen.PLAN_COLS.index("initiative") + 1
    ws.cell(row=2, column=init_col, value="SA")
    ws.cell(row=3, column=init_col, value="INTEL")
    wb.save(template)

    _cols, records, warnings = gen.generate(template, output)
    assert records and not warnings
    seen_with, seen_without = False, False
    for rec in records:
        d = nd.decode_campaign_name(rec["Campaign Name"])
        assert d.kind == "campaign"
        assert d.initiative == rec["initiative"]          # round-trips, blank or set
        assert d.market == rec["market"]
        assert d.channel == rec["channel"]                # name carries the raw plan channel
        assert d.objective == rec["objective"]
        assert d.audience_type == rec["audience_type"]
        seen_with = seen_with or d.initiative != ""
        seen_without = seen_without or d.initiative == ""
    assert seen_with and seen_without


def test_blank_is_campaign_level_not_unparsed():
    for blank in ("", "  ", None, float("nan")):
        d = nd.decode_name(blank)
        assert d.kind == "blank"
        assert d.audience_type == ""    # a campaign-level row is not a naming failure


def test_decode_series_and_unparsed_rate():
    names = pd.Series(["PROSPECT_LAL-1PCT_FEED", "garbage name!", "BRANDHERO_VID_9x16",
                       "", "RETARGET_SITE-90D_FEED"])
    decoded = nd.decode_series(names)
    assert list(decoded.columns) == list(nd.FIELD_COLUMNS)
    assert decoded.loc[0, "audience_type"] == "PROSPECT"
    assert decoded.loc[1, "audience_type"] == nd.UNPARSED
    assert decoded.loc[2, "creative"] == "BRANDHERO"      # ad name: audience stays empty
    assert decoded.loc[3, "audience_type"] == ""          # campaign-level row

    df = pd.DataFrame({"ad_group": names, "audience_type": decoded["audience_type"]})
    # denominator = the 4 ad-level rows (blank excluded); 1 unparsed among them
    assert nd.unparsed_rate(df) == pytest.approx(0.25)
    assert nd.unparsed_rate(pd.DataFrame({"ad_group": [""], "audience_type": [""]})) == 0.0


def test_crosswalk_overrides_take_precedence():
    """A curated crosswalk maps a non-conforming name to real fields, before the grammar."""
    ov = {nd.norm_key("Veterans - Feed"): {"audience_type": "PROSPECT",
                                            "audience_detail": "VETERANS"},
          nd.norm_key("MissionHero video 1x1"): {"creative": "MISSIONHERO",
                                                  "creative_format": "VID"}}
    # matches ignore case/whitespace
    f = nd.decode_fields("veterans   -   feed", overrides=ov)
    assert f["audience_type"] == "PROSPECT" and f["audience_detail"] == "VETERANS"
    f2 = nd.decode_fields("MissionHero video 1x1", overrides=ov)
    assert f2["creative"] == "MISSIONHERO" and f2["creative_format"] == "VID"
    # a name not in the crosswalk still falls to the grammar / unparsed
    assert nd.decode_fields("Advantage+ broad TEST", overrides=ov)["audience_type"] == nd.UNPARSED
    # grammar still wins for conforming names when no override matches
    assert nd.decode_fields("PROSPECT_LAL-1PCT_FEED", overrides=ov)["audience_type"] == "PROSPECT"
