"""Export the cleaned, organized data to a multi-sheet Excel workbook.

A hand-off / audit deliverable built from the durable store (``history.parquet``) and
the weekly modeling table. Sheets:

  Name mapping   every raw ad-set/creative name -> its decoded audience/creative fields
                 and HOW it was resolved (convention | crosswalk | unparsed), with spend
                 — so you can see exactly what the naming clean-up changed.
  Cleaned daily  the canonical daily rows (raw name kept, decoded fields alongside).
  Campaigns      per channel x campaign rollup (claimed conv + GA4 key events).
  Audiences      decoded-audience rollup with cost-per and spend/claim shares.
  Weekly         the national weekly modeling table (channel_weekly_metrics.csv).

Usage:  python scripts/export_clean.py [--out outputs/cleaned_export.xlsx]
"""
from __future__ import annotations

import argparse
import sys
from pathlib import Path

import pandas as pd

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / "src"))

from advanced_reporting.dashboard import drilldown            # noqa: E402
from advanced_reporting.ingestion import naming_decode        # noqa: E402
from advanced_reporting.utils import load_naming_overrides    # noqa: E402

_DAILY_COLS = ["date", "source", "channel", "campaign", "ad_group",
               "audience_type", "audience_detail", "creative", "creative_format", "geo",
               "spend", "impressions", "clicks", "video_views", "conversions",
               "platform_revenue", "sessions", "engaged_sessions", "page_views",
               "key_events"]


def _resolution(name: str, overrides: dict) -> str:
    """How this name got its fields: convention (grammar) | crosswalk | unparsed."""
    if naming_decode.decode_name(name).kind in ("ad_set", "ad"):
        return "convention"
    if naming_decode.norm_key(name) in overrides:
        return "crosswalk"
    return "unparsed"


def _name_mapping(hist: pd.DataFrame, overrides: dict) -> pd.DataFrame:
    ad = hist[hist["ad_group"].fillna("").astype(str).str.strip() != ""]
    per = (ad.groupby(["channel", "ad_group", "audience_type", "audience_detail",
                       "creative", "creative_format"], as_index=False)
             .agg(spend=("spend", "sum"), rows=("ad_group", "size")))
    per["resolved_by"] = per["ad_group"].map(lambda n: _resolution(n, overrides))
    per = per.rename(columns={"ad_group": "raw_name"})
    order = {"convention": 0, "crosswalk": 1, "unparsed": 2}
    per["_o"] = per["resolved_by"].map(order)
    return (per.sort_values(["_o", "spend"], ascending=[True, False])
               .drop(columns="_o")[["raw_name", "channel", "resolved_by", "audience_type",
                                    "audience_detail", "creative", "creative_format",
                                    "spend", "rows"]])


def _autofmt(ws, df: pd.DataFrame) -> None:
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter
    for j, col in enumerate(df.columns, start=1):
        ws.cell(row=1, column=j).font = Font(bold=True)
        width = max(len(str(col)), *(df[col].astype(str).str.len().head(200).tolist() or [0]))
        ws.column_dimensions[get_column_letter(j)].width = min(max(width + 2, 10), 42)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def main(argv=None) -> None:
    ap = argparse.ArgumentParser(description="Export cleaned data to an Excel workbook.")
    ap.add_argument("--out", default=str(ROOT / "outputs" / "cleaned_export.xlsx"))
    args = ap.parse_args(argv)

    hist_path = ROOT / "data" / "processed" / "history.parquet"
    if not hist_path.exists():
        print("No history.parquet — run ingest + run_pipeline first.")
        return
    hist = pd.read_parquet(hist_path)
    overrides = load_naming_overrides()

    sheets: dict[str, pd.DataFrame] = {}
    sheets["Name mapping"] = _name_mapping(hist, overrides)
    daily = hist[[c for c in _DAILY_COLS if c in hist.columns]].sort_values(
        ["date", "channel", "campaign", "ad_group"])
    sheets["Cleaned daily"] = daily
    sheets["Campaigns"] = drilldown.campaign_table(hist)
    aud = drilldown.audience_summary(hist)
    if not aud.empty:
        sheets["Audiences"] = aud
    weekly_f = ROOT / "data" / "processed" / "channel_weekly_metrics.csv"
    if weekly_f.exists():
        sheets["Weekly"] = pd.read_csv(weekly_f)

    out = Path(args.out)
    out.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(out, engine="openpyxl") as xw:
        for name, df in sheets.items():
            df.to_excel(xw, sheet_name=name, index=False)
            _autofmt(xw.sheets[name], df)

    nm = sheets["Name mapping"]
    by = nm["resolved_by"].value_counts().to_dict()
    print(f"Wrote {out.relative_to(ROOT)}  ({len(sheets)} sheets)")
    for name, df in sheets.items():
        print(f"  {name:<14} {len(df):>6,} rows")
    print(f"  name resolution: {by}")


if __name__ == "__main__":
    main()
