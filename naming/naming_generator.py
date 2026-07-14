"""Naming-convention generator for ad/marketing campaigns (Plan-row model).

Commands:
  template <out.xlsx>             build a pre-filled input template to fill in
  generate <in.xlsx> <out.xlsx>  read a filled template -> names + UTMs + validation

INPUT MODEL — one Plan sheet, one row per real line item you intend to run.
Each cell may hold a single value OR a comma-list. The tool expands the cross-product
*within a row only* (so placements/sizes fan out per line) and SUMS across rows. This
matches reality (creatives/sizes are nested under an audience) and avoids the global
cross-product explosion.

A naming convention is still a grammar: the Scheme composes each level (Campaign / Ad Set
/ Ad) from ordered fields. Generating from a fixed grammar is what lets the data be
DECODED later with the same scheme.
"""
from __future__ import annotations
import sys
import itertools
import re
from urllib.parse import urlencode
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

FONT = "Arial"
HEAD_FILL = PatternFill("solid", fgColor="1F4E78")
HEAD_FONT = Font(name=FONT, bold=True, color="FFFFFF", size=11)
NOTE_FONT = Font(name=FONT, italic=True, color="555555", size=10)
WARN_FILL = PatternFill("solid", fgColor="FCE4D6")
LIST_FILL = PatternFill("solid", fgColor="E2EFDA")
thin = Side(style="thin", color="BFBFBF")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

PLAN_COLS = ["market", "channel", "objective", "audience_type", "audience_detail",
             "creative", "format", "size", "placement", "version", "initiative"]
# ``initiative`` is an OPTIONAL trailing campaign segment (e.g. a career path / product
# line): US_META_CONVERT_PROSPECT_SA. Blank -> the campaign name is the classic 4-segment
# form (market_channel_objective_audience_type), so every pre-initiative plan still
# generates and decodes unchanged. Decoded by ``ingestion/naming_decode.decode_campaign_name``.
# Example plan: lists live in one cell (format/size/placement) and expand within the row.
EXAMPLE_PLAN = [
    ["US", "META",   "CONVERT", "PROSPECT", "INT-FITNESS",     "SUMMERSALE", "VID",         "9x16",        "FEED, REELS", "V1", ""],
    ["US", "META",   "CONVERT", "PROSPECT", "LAL-1PCT",        "BRANDHERO",  "STATIC",      "1x1, 4x5",    "FEED",        "V1", ""],
    ["US", "TIKTOK", "CONVERT", "RETARGET", "CART-ABANDON",    "SUMMERSALE", "VID",         "9x16",        "FEED",        "V1", ""],
    ["US", "TIKTOK", "CONVERT", "RETARGET", "PDP-VIEW-30D",    "BRANDHERO",  "VID, STATIC", "9x16",        "REELS",       "V1", ""],
]
DEFAULT_SCHEME = [
    ("Campaign", "market, channel, objective, audience_type, initiative", "_"),
    ("Ad Set",   "audience_type, audience_detail, placement", "_"),
    ("Ad",       "creative, format, size, version",           "_"),
]
SETTINGS_DEFAULTS = [
    ("case", "asis", "asis | upper | lower  -- how to cap tokens in the generated names"),
    ("max_name_length", "100", "Flag any generated name longer than this many characters"),
    ("max_combinations", "20000", "Abort if the total would exceed this (a guardrail)"),
    ("landing_url", "https://www.example.com/summer-sale", "Base URL the UTM parameters are appended to"),
    ("utm_medium", "paid", "Value used for utm_medium on every generated URL"),
]
CHANNEL_CANON = {
    "META": "meta", "FB": "meta", "IG": "meta",
    "TIKTOK": "tiktok", "TT": "tiktok",
    "GOOGLE": "google_search", "GSEARCH": "google_search", "SEARCH": "google_search",
    "PMAX": "google_pmax", "LI": "linkedin", "LINKEDIN": "linkedin",
}
PREFERRED_FIELD_ORDER = ["market", "channel", "objective", "audience_type", "initiative",
                         "audience_detail", "audience", "creative", "format", "size",
                         "placement", "version"]
NAME_RE = re.compile(r"^[A-Za-z0-9_\-]+$")
NAME_COLS = ["Campaign Name", "Ad Set Name", "Ad Name",
             "Landing URL (with UTMs)", "Ad Set Key (Campaign|Set)"]


def _style_header(ws, row, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEAD_FILL; cell.font = HEAD_FONT
        cell.alignment = Alignment(horizontal="left", vertical="center"); cell.border = BORDER


def _autosize(ws, maxw=60):
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        width = max((len(str(c.value)) for c in col if c.value is not None), default=10)
        ws.column_dimensions[letter].width = min(max(width + 2, 10), maxw)


def _ordered_fields(fields):
    return [f for f in PREFERRED_FIELD_ORDER if f in fields] + [f for f in fields if f not in PREFERRED_FIELD_ORDER]


# ---------------------------------------------------------------- template
def build_template(path):
    wb = Workbook(); wb.remove(wb.active)

    ws = wb.create_sheet("README")
    lines = [
        ("Campaign Naming Convention - Generator Template (Plan model)", 13),
        ("", 0),
        ("HOW TO USE", 11),
        ("1. Plan sheet: one ROW per real line item you intend to run.", 0),
        ("   A cell can be a single value OR a comma-list. Comma-lists expand WITHIN that row", 0),
        ("   only (e.g. placement = 'FEED, REELS' -> 2 ads). Totals SUM across rows, so there", 0),
        ("   is no global cross-product explosion -- you list only the combinations you want.", 0),
        ("2. Scheme sheet: how each level's name is assembled from the fields, in order.", 0),
        ("3. Settings sheet: casing, length guardrails, landing URL for the UTMs.", 0),
        ("4. Save, then:  python naming_generator.py generate <this file> output.xlsx", 0),
        ("", 0),
        ("ROW VOLUME = product of the comma-list lengths in that row. Single-value cells = x1.", 10),
        ("", 0),
        ("WHY THIS SHAPE", 11),
        ("Generating from a fixed grammar means the exported data can later be DECODED back", 0),
        ("into these fields. Channel codes map to canonical channels and utm_source, so the", 0),
        ("UTMs line up with analytics (e.g. GA4) on the way back in.", 0),
    ]
    for i, (text, sz) in enumerate(lines, start=1):
        ws.cell(row=i, column=1, value=text).font = Font(name=FONT, bold=sz >= 11, italic=sz == 10, size=sz or 10)
    ws.column_dimensions["A"].width = 100

    ws = wb.create_sheet("Plan")
    ws.append(PLAN_COLS); _style_header(ws, 1, len(PLAN_COLS))
    list_cols = {PLAN_COLS.index("format"), PLAN_COLS.index("size"), PLAN_COLS.index("placement")}
    for r in EXAMPLE_PLAN:
        ws.append(r)
        for j in range(len(PLAN_COLS)):
            cell = ws.cell(row=ws.max_row, column=j + 1); cell.font = Font(name=FONT); cell.border = BORDER
            if j in list_cols and "," in str(r[j]):
                cell.fill = LIST_FILL
    ws.freeze_panes = "A2"; _autosize(ws, 26)

    ws = wb.create_sheet("Scheme")
    ws.append(["Level", "Fields (ordered, comma-separated)", "Delimiter"])
    for level, fields, delim in DEFAULT_SCHEME:
        ws.append([level, fields, delim])
    _style_header(ws, 1, 3)
    for r in range(2, 2 + len(DEFAULT_SCHEME)):
        for c in range(1, 4):
            ws.cell(row=r, column=c).font = Font(name=FONT); ws.cell(row=r, column=c).border = BORDER
    ws.column_dimensions["A"].width = 14; ws.column_dimensions["B"].width = 46; ws.column_dimensions["C"].width = 11

    ws = wb.create_sheet("Settings")
    ws.append(["Setting", "Value", "Notes"])
    for k, v, note in SETTINGS_DEFAULTS:
        ws.append([k, v, note])
    _style_header(ws, 1, 3)
    for r in range(2, 2 + len(SETTINGS_DEFAULTS)):
        ws.cell(row=r, column=1).font = Font(name=FONT, bold=True)
        ws.cell(row=r, column=2).font = Font(name=FONT); ws.cell(row=r, column=2).fill = PatternFill("solid", fgColor="FFF2CC")
        ws.cell(row=r, column=3).font = NOTE_FONT
        for c in range(1, 4): ws.cell(row=r, column=c).border = BORDER
    ws.column_dimensions["A"].width = 18; ws.column_dimensions["B"].width = 40; ws.column_dimensions["C"].width = 64

    wb.save(path)
    print(f"template -> {path}")


# ---------------------------------------------------------------- read
def _read_template(path):
    wb = load_workbook(path, data_only=True)
    settings = {k: v for k, v, *_ in SETTINGS_DEFAULTS}
    if "Settings" in wb.sheetnames:
        for row in wb["Settings"].iter_rows(min_row=2, values_only=True):
            if row and row[0]:
                settings[str(row[0]).strip()] = "" if row[1] is None else str(row[1]).strip()
    scheme = []
    for row in wb["Scheme"].iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        fields = [f.strip() for f in str(row[1]).replace(";", ",").split(",") if f.strip()]
        delim = "_" if row[2] in (None, "") else str(row[2])
        scheme.append((str(row[0]).strip(), fields, delim))
    pws = wb["Plan"]; cols = [str(c.value).strip() for c in pws[1] if c.value]
    plan = []
    for row in pws.iter_rows(min_row=2, values_only=True):
        rec = {cols[j]: ("" if v is None else str(v).strip()) for j, v in enumerate(row) if j < len(cols)}
        if sum(1 for v in rec.values() if v) >= 2:
            plan.append(rec)
    return cols, plan, scheme, settings


def _case(tok, mode):
    return tok.upper() if mode == "upper" else tok.lower() if mode == "lower" else tok


def _opts(cell):
    return [x.strip() for x in str(cell).split(",") if x.strip()] or [""]


# ---------------------------------------------------------------- generate
def generate(in_path, out_path):
    plan_cols, plan, scheme, settings = _read_template(in_path)
    case = settings.get("case", "asis")
    max_len = int(float(settings.get("max_name_length", 100)))
    max_combos = int(float(settings.get("max_combinations", 20000)))
    landing = settings.get("landing_url", "").strip()
    utm_medium = settings.get("utm_medium", "paid")

    used = []
    for _, fields, _d in scheme:
        for f in fields:
            if f not in used:
                used.append(f)
    missing = [f for f in used if f not in plan_cols]
    if missing:
        sys.exit(f"ERROR: Scheme uses fields not present as Plan columns: {missing}")

    total = sum(_row_count(row, plan_cols) for row in plan)
    if total > max_combos:
        sys.exit(f"ERROR: {total} combinations exceeds max_combinations={max_combos}. "
                 f"Trim comma-lists or raise the cap in Settings.")

    records, warnings, seen = [], [], set()
    for row in plan:
        per_col = [_opts(row.get(c, "")) for c in plan_cols]
        for combo in itertools.product(*per_col):
            vals = {c: _case(v, case) for c, v in zip(plan_cols, combo)}
            names = {lvl: delim.join(vals[f] for f in fields if vals.get(f, "").strip())
                     for lvl, fields, delim in scheme}
            canon = CHANNEL_CANON.get(vals.get("channel", "").upper(), vals.get("channel", "").lower())
            term = "-".join(x for x in (vals.get("audience_type", ""),
                                        vals.get("audience_detail", "")) if x).lower()
            utms = {"utm_source": canon, "utm_medium": utm_medium,
                    "utm_campaign": names.get("Campaign", "").lower(),
                    "utm_content": names.get("Ad", "").lower()}
            if term:
                utms["utm_term"] = term
            url = f"{landing}?{urlencode(utms)}" if landing else f"?{urlencode(utms)}"
            rec = dict(vals)
            rec["canonical_channel"] = canon
            rec["Campaign Name"] = names.get("Campaign", "")
            rec["Ad Set Name"] = names.get("Ad Set", "")
            rec["Ad Name"] = names.get("Ad", "")
            rec["Landing URL (with UTMs)"] = url
            rec["Ad Set Key (Campaign|Set)"] = f'{names.get("Campaign", "")}|{names.get("Ad Set", "")}'
            records.append(rec)
            for lvl in ("Campaign", "Ad Set", "Ad"):
                nm = names.get(lvl, "")
                if nm and not NAME_RE.match(nm):
                    warnings.append((nm, lvl, "illegal characters (allowed: A-Z a-z 0-9 _ -)"))
                if len(nm) > max_len:
                    warnings.append((nm, lvl, f"exceeds max length {max_len} ({len(nm)} chars)"))
            key = rec["Ad Set Key (Campaign|Set)"] + "|" + rec["Ad Name"]
            if key in seen:
                warnings.append((rec["Ad Name"], "Ad", "duplicate Campaign|AdSet|Ad combination"))
            seen.add(key)

    cols = []
    for f in _ordered_fields(plan_cols):
        cols.append(f)
        if f == "channel":
            cols.append("canonical_channel")
    cols += NAME_COLS
    _write_output(out_path, cols, records, warnings)
    print(f"generate -> {out_path}  ({len(records)} ads from {len(plan)} plan rows, {len(warnings)} warnings)")
    return cols, records, warnings


def _row_count(row, cols):
    n = 1
    for c in cols:
        n *= len(_opts(row.get(c, "")))
    return n


def _write_output(path, cols, records, warnings):
    wb = Workbook(); wb.remove(wb.active)
    last = len(records) + 1
    L = {c: get_column_letter(i + 1) for i, c in enumerate(cols)}

    ws = wb.create_sheet("Trafficking Sheet")
    ws.append(cols); _style_header(ws, 1, len(cols))
    for rec in records:
        ws.append([rec.get(c, "") for c in cols])
    for i in range(2, last + 1):
        for j in range(1, len(cols) + 1):
            cell = ws.cell(row=i, column=j); cell.font = Font(name=FONT, size=10); cell.border = BORDER
    ws.freeze_panes = "A2"; _autosize(ws, 70)

    ws = wb.create_sheet("Validation")
    ws.append(["Name", "Level", "Issue"]); _style_header(ws, 1, 3)
    if warnings:
        for w in warnings:
            ws.append(list(w))
            for j in range(1, 4):
                ws.cell(row=ws.max_row, column=j).fill = WARN_FILL
                ws.cell(row=ws.max_row, column=j).font = Font(name=FONT, size=10)
    else:
        ws.append(["No issues found.", "", ""])
        ws.cell(row=2, column=1).font = Font(name=FONT, size=10, color="375623")
    _autosize(ws, 70)

    ws = wb.create_sheet("Summary")
    tr = "'Trafficking Sheet'"
    ws.append(["Metric", "Value"]); _style_header(ws, 1, 2)
    ws.append(["Total ads", f"=COUNTA({tr}!{L['Ad Name']}2:{L['Ad Name']}{last})"])
    ws.append(["Unique campaigns",
               f"=SUMPRODUCT(1/COUNTIF({tr}!{L['Campaign Name']}2:{L['Campaign Name']}{last},"
               f"{tr}!{L['Campaign Name']}2:{L['Campaign Name']}{last}))"])
    k = L["Ad Set Key (Campaign|Set)"]
    ws.append(["Unique ad sets (campaign+set)",
               f"=SUMPRODUCT(1/COUNTIF({tr}!{k}2:{k}{last},{tr}!{k}2:{k}{last}))"])
    ws.append([])
    ws.append(["Ads per channel", ""]); _style_header(ws, ws.max_row, 2)
    cc = L["canonical_channel"]
    for canon in sorted({r["canonical_channel"] for r in records}):
        ws.append([canon, f'=COUNTIF({tr}!{cc}2:{cc}{last},"{canon}")'])
    for i in range(2, ws.max_row + 1):
        for j in (1, 2):
            cell = ws.cell(row=i, column=j)
            if not (cell.font and cell.font.color):
                cell.font = Font(name=FONT, size=10)
            cell.border = BORDER
    ws.column_dimensions["A"].width = 30; ws.column_dimensions["B"].width = 16
    wb.save(path)


if __name__ == "__main__":
    if len(sys.argv) >= 3 and sys.argv[1] == "template":
        build_template(sys.argv[2])
    elif len(sys.argv) >= 4 and sys.argv[1] == "generate":
        generate(sys.argv[2], sys.argv[3])
    else:
        print("usage:\n  naming_generator.py template <out.xlsx>\n"
              "  naming_generator.py generate <in.xlsx> <out.xlsx>")
        sys.exit(1)
